"""
src/screener/export_peer_comparison.py — Nifty 100 Analytics
Sprint 3 / Day 20: output/peer_comparison.xlsx — 11 sheets, one per peer
group. Each sheet: company_id, company_name, + metric columns + percentile
rank per metric, percentile color-coded, benchmark row highlighted gold,
summary row with peer group median.
"""
from __future__ import annotations

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "analytics"))
sys.path.insert(0, os.path.dirname(__file__))
from universe import build_universe  # noqa: E402
from peer import compute_peer_percentiles, METRICS  # noqa: E402

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "peer_comparison.xlsx")

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


def _pctile_fill(pctile):
    if pctile is None or pd.isna(pctile):
        return None
    if pctile >= 0.75:
        return GREEN
    if pctile >= 0.25:
        return YELLOW
    return RED


def write_group_sheet(wb, group_name: str, percentiles: pd.DataFrame,
                       universe: pd.DataFrame, benchmark_ids: set):
    ws = wb.create_sheet(title=group_name[:31])

    metric_names = list(METRICS.keys())
    headers = ["company_id", "company_name"]
    for m in metric_names:
        headers += [m, f"{m} %ile"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    group_df = percentiles[percentiles["peer_group_name"] == group_name]
    company_ids = group_df["company_id"].unique().tolist()
    company_names = universe.set_index("company_id")["company_name"].to_dict()

    pivot_value = group_df.pivot_table(index="company_id", columns="metric", values="value", aggfunc="first")
    pivot_pct = group_df.pivot_table(index="company_id", columns="metric", values="percentile_rank", aggfunc="first")

    row_num = 2
    for cid in company_ids:
        row = [cid, company_names.get(cid, cid)]
        for m in metric_names:
            val = pivot_value.loc[cid, m] if cid in pivot_value.index and m in pivot_value.columns else None
            pct = pivot_pct.loc[cid, m] if cid in pivot_pct.index and m in pivot_pct.columns else None
            row += [round(val, 2) if pd.notna(val) else None, round(pct, 2) if pd.notna(pct) else None]
        ws.append(row)

        if cid in benchmark_ids:
            for c in range(1, len(row) + 1):
                ws.cell(row=row_num, column=c).fill = GOLD

        for i, m in enumerate(metric_names):
            pct_col = 2 + (i * 2) + 2  # 1-indexed column of the %ile column for metric i
            pct_val = pivot_pct.loc[cid, m] if cid in pivot_pct.index and m in pivot_pct.columns else None
            fill = _pctile_fill(pct_val)
            if fill and cid not in benchmark_ids:
                ws.cell(row=row_num, column=pct_col).fill = fill
        row_num += 1

    # Summary row: peer group median per metric
    summary_row = ["MEDIAN", ""]
    for m in metric_names:
        med = pivot_value[m].median() if m in pivot_value.columns else None
        summary_row += [round(med, 2) if pd.notna(med) else None, None]
    ws.append(summary_row)
    for c in range(1, len(summary_row) + 1):
        ws.cell(row=row_num, column=c).font = BOLD

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


def run():
    import sqlite3
    conn = sqlite3.connect(os.path.join(BASE_DIR, "nifty100.db"))
    universe = build_universe(conn)
    peer_groups = pd.read_sql("SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn)
    conn.close()

    percentiles = compute_peer_percentiles(universe, peer_groups)
    benchmark_ids = set(peer_groups.loc[peer_groups["is_benchmark"] == 1, "company_id"])

    wb = Workbook()
    wb.remove(wb.active)

    for group_name in sorted(peer_groups["peer_group_name"].unique()):
        write_group_sheet(wb, group_name, percentiles, universe, benchmark_ids)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH} ({len(wb.sheetnames)} sheets)")
    return wb.sheetnames


if __name__ == "__main__":
    run()
