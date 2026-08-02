"""
src/screener/export_screener.py — Nifty 100 Analytics
Sprint 3 / Day 17: generate output/screener_output.xlsx — one sheet per
preset, 20 KPI columns, sorted by composite_quality_score descending,
with green/red fill cells showing pass/fail against that preset's own
thresholds.
"""
from __future__ import annotations

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "analytics"))
sys.path.insert(0, os.path.dirname(__file__))
from universe import build_universe  # noqa: E402
from engine import load_config, run_preset, PRESET_NAMES  # noqa: E402

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "screener_output.xlsx")

DISPLAY_COLUMNS = [
    "company_id", "company_name", "broad_sector", "composite_quality_score",
    "return_on_equity_pct", "roce_pct", "net_profit_margin_pct", "operating_profit_margin_pct",
    "debt_to_equity", "interest_coverage", "icr_label",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "free_cash_flow_cr", "cfo_quality_label", "fcf_conversion_pct",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "market_cap_crore",
]

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _passes_metric(row, metric_name, threshold, config) -> bool | None:
    spec = config["metrics"].get(metric_name)
    if spec is None:
        return None
    col, direction = spec["column"], spec["direction"]
    val = row.get(col)
    if col == "debt_to_equity" and row.get("broad_sector") == "Financials":
        return True
    if col == "interest_coverage" and row.get("icr_label") == "Debt Free":
        return True
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return val >= threshold if direction == "min" else val <= threshold


def write_preset_sheet(wb, preset_name, universe, config):
    preset = config["presets"][preset_name]
    result = run_preset(preset_name, universe, config)
    result = result[[c for c in DISPLAY_COLUMNS if c in result.columns]].copy()

    ws = wb.create_sheet(title=preset["label"][:31])
    ws.append(list(result.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for _, row in result.iterrows():
        ws.append(list(row.values))

    # Color-code cells for columns that map to a threshold in this preset's filters
    filters = preset["filters"]
    col_index = {name: i + 1 for i, name in enumerate(result.columns)}
    for metric_name, threshold in filters.items():
        spec = config["metrics"].get(metric_name)
        if spec is None or spec["column"] not in col_index:
            continue
        col_letter = get_column_letter(col_index[spec["column"]])
        for r, (_, row) in enumerate(result.iterrows(), start=2):
            passed = _passes_metric(row, metric_name, threshold, config)
            if passed is True:
                ws[f"{col_letter}{r}"].fill = GREEN
            elif passed is False:
                ws[f"{col_letter}{r}"].fill = RED

    for i, col in enumerate(result.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    return len(result)


def run():
    universe = build_universe()
    config = load_config()

    wb = Workbook()
    wb.remove(wb.active)

    counts = {}
    for preset_name in PRESET_NAMES:
        n = write_preset_sheet(wb, preset_name, universe, config)
        counts[preset_name] = n

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH}")
    for name, n in counts.items():
        flag = "" if 5 <= n <= 50 else "  <-- outside 5-50 target range"
        print(f"  {name}: {n} companies{flag}")
    return counts


if __name__ == "__main__":
    run()
